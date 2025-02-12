import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

def generate_excel(data, filename="generated_report.xlsx"):
    # Crear un DataFrame con los datos
    df = pd.DataFrame(data)
    
    # Guardar en un archivo Excel
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Reporte', index=False)
        
        # Obtener el libro y la hoja activa
        workbook = writer.book
        sheet = writer.sheets['Reporte']
        
        # Aplicar estilos a las celdas
        bold_font = Font(bold=True)
        fill_color = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Aplicar estilos a la cabecera
        for col_num, column_title in enumerate(df.columns, 1):
            cell = sheet[f'{get_column_letter(col_num)}1']
            cell.font = bold_font
            cell.fill = fill_color
            cell.border = border_style
            
        # Ajustar el ancho de las columnas
        for col_num, column_title in enumerate(df.columns, 1):
            sheet.column_dimensions[get_column_letter(col_num)].width = 15
            
        # Insertar una fórmula en una nueva columna (Ejemplo: Suma de la columna B)
        last_row = len(df) + 1
        sum_cell = f'B{last_row+1}'
        sheet[sum_cell] = f'=SUM(B2:B{last_row})'
        sheet[sum_cell].font = bold_font
        
        # Crear un gráfico de barras
        chart = BarChart()
        chart.title = "Análisis de Datos"
        data_range = Reference(sheet, min_col=2, min_row=1, max_row=last_row, max_col=2)
        categories = Reference(sheet, min_col=1, min_row=2, max_row=last_row)
        chart.add_data(data_range, titles_from_data=True)
        chart.set_categories(categories)
        sheet.add_chart(chart, "D5")
    
    return filename

# Datos de prueba
data = {
    "Categoría": ["Ventas", "Marketing", "Operaciones", "Finanzas"],
    "Ingresos": [50000, 30000, 40000, 45000]
}

# Generar el archivo Excel
generate_excel(data)
